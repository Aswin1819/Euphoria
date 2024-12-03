from django.shortcuts import render,redirect,get_object_or_404
from django.contrib.auth import login,logout,authenticate
from django.contrib import messages
from django.contrib.auth import get_user_model
from .models import EuphoUser,Category,Products,Images,Brand,Variant,Order,OrderItem,Coupon,UserCoupon,Offer
from .forms import ProductForm, VariantForm,CouponForm,OfferForm
from django.forms import modelformset_factory
from django.db.models import Q,Sum,Count,F,DecimalField
from django.utils.timezone import now, make_aware
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.views.decorators.cache import never_cache
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from userapp.utils import refund_to_wallet
import json
from django.views.decorators.http import require_POST
from django.db.models import F
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth, TruncYear
from django.http import JsonResponse
from .utils import superuser_required
#imagecropping modules
from django.core.files.base import ContentFile
import base64
import re

# Create your views here.

                                #########           #########
                                #########ADMIN LOGIN#########
                                #########           #########
@never_cache
def adminLogin(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect(adminCustomers)
        return render(request,'adminlogin.html')    
    return render(request,'adminlogin.html')


def checkAdmin(request):
        if request.method == 'POST':
            email = request.POST.get('email')
            password = request.POST.get('password')
            
            if not email or not password:
                messages.warning(request,"Email and password are required")
                return render(request,'adminlogin.html')
            
            User = get_user_model()
            
            user = authenticate(request,email=email,password=password)
            if user is not None:
                login(request,user)
                if user.is_staff:
                    return redirect(adminLogin)
                else:
                    messages.warning(request,'You do not have permission to admin site')
                    return render(request,'adminlogin.html')
            else:
                messages.warning(request,"Invalid Email or Password")
                return render(request,'adminlogin.html')
        else:
            messages.warning(request,"Invalid Credentials!")
            return render(request,'adminlogin.html')
    
        
@never_cache
def adminLogout(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            logout(request)
        return redirect(adminLogin)
    return redirect(adminLogin)
                
        
        
                                #########           #########
                                #########ADMIN CUSTOMERS#####
                                #########           #########    
        
        
@superuser_required
def adminCustomers(request):   
    user_list = EuphoUser.objects.all().order_by('-updated_at')
        
    paginator = Paginator(user_list,5)
    page_number = request.GET.get('page')
    user = paginator.get_page(page_number)
        
    return render(request,'admincustomers.html',{'users':user})
    


User = get_user_model()
@never_cache
def blockUser(request,id):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            user = get_object_or_404(User,id=id)
            user.is_active = not user.is_active
            user.save()
            return redirect(adminCustomers)
        return render(request,'adminlogin.html')
    return render(request,'adminlogin.html')
            

@never_cache
@superuser_required
def customerSearch(request):
    if request.method=='POST':
        search = request.POST.get('search')
        if search is not None:
            user = User.objects.filter(username__icontains=search).order_by('username')
        else:
            user = User.objects.all()
        return render(request,'adminCustomers.html',{'users':user})


                    #########           #########
                    #########ADMIN CATEGORY######
                    #########           #########


@superuser_required
def adminCategory(request):
    if request.user.is_superuser:
        category_list = Category.objects.all().order_by('-updated_date')
        
        paginator = Paginator(category_list,5)
        page_number = request.GET.get('page')
        category = paginator.get_page(page_number)
        
        return render(request,'admincategory.html',{'items':category})
    return redirect(adminLogin)


@never_cache
@superuser_required
def addCategory(request):
    if request.user.is_superuser:
        if request.method == "POST":
            category = request.POST.get('category')
            if Category.objects.filter(name__iexact=category).exists():
                messages.warning(request,"Category name already exists")
                return redirect(adminCategory)
            if not category:
                messages.warning(request,"name cant be empty")
                return redirect(adminCategory)
            if not re.match("^[A-Za-z ]*$", category) or not category.strip():
                messages.warning(request, "Name must contain only letters and cannot be empty.")
                return redirect(adminCategory)
            if len(category)>20:
                messages.warning(request,"Name must be less than 20characters")
                return redirect(adminCategory)
            
            new_category = Category.objects.create(name=category)
            new_category.save()
            print("category added successfully")
            return redirect(adminCategory)
    return redirect(adminLogin)

@never_cache
@superuser_required
def removeCategory(request,id):
    if request.user.is_superuser:
        if request.user.is_authenticated:            
            item = Category.objects.get(id=id)
            item.is_active = not item.is_active
            item.save()
            return redirect(adminCategory)
        
    return redirect(adminLogin)


@never_cache
@superuser_required
def editCategory(request,id):
    if request.user.is_superuser:
        if request.method=='POST':
            new_name = request.POST.get('category')
            if Category.objects.filter(name__iexact=new_name).exclude(id=id).exists():
                messages.warning(request,"Category name already exists")
                return redirect(adminCategory)
            if not new_name:
                messages.warning(request,"name cant be empty")
                return redirect(adminCategory)
            if not re.match("^[A-Za-z ]*$", new_name) or not new_name.strip():
                messages.warning(request, "Name must contain only letters and cannot be empty.")
                return redirect(adminCategory)
            if len(new_name)>20:
                messages.warning(request,"Name must be less than 20characters")
                return redirect(adminCategory)
            
            item = Category.objects.get(id=id)
            item.name = new_name
            item.save()
            messages.success(request,'Category edited successfully')
            return redirect(adminCategory)
    return redirect(adminLogin)


@never_cache
@superuser_required
def categorySearch(request):
    if request.user.is_superuser:
        if request.method=='POST':
            search = request.POST.get('search')
            if search is not None:
                item = Category.objects.filter(name__icontains=search).order_by('name')
            else:
                item = Category.objects.all()
            return render(request,'admincategory.html',{'items':item})



                            #########           #########
                            #########ADMIN PRODUCTS######
                            #########           #########



@never_cache
@superuser_required
def adminProducts(request):
    if request.user.is_superuser:
        product_list = Products.objects.all().prefetch_related('variants').order_by('-updated_at')
        category = Category.objects.all()
        
        paginator = Paginator(product_list,5)
        page_number = request.GET.get('page')
        product = paginator.get_page(page_number)
        
        return render(request,'adminProducts.html',{'products':product,'categories':category})
    return redirect(adminLogin)


@never_cache
@superuser_required
def addProducts(request):
    if request.user.is_superuser:
        VariantFormSet = modelformset_factory(Variant,form=VariantForm,extra=1)
        
        if request.method == 'POST':
            product_form = ProductForm(request.POST) #request.FILES changed
            variant_formset = VariantFormSet(request.POST)
            
            image_files = [
                    request.POST.get('image1'),
                    request.POST.get('image2'),
                    request.POST.get('image3'),
                    request.POST.get('image4'),
                 ]
            
            if product_form.is_valid() and variant_formset.is_valid():
                print("form and variants are valid ")
                product = product_form.save()
                
                for image_data in image_files:
                    if image_data:  
                        try:
                            format, imgstr = image_data.split(';base64,')
                            ext = format.split('/')[-1]  
                            image_file = ContentFile(base64.b64decode(imgstr), name=f'product_image.{ext}')
                            Images.objects.create(images=image_file, product=product)
                        except Exception as e:
                            print(f"Error processing image: {e}")
                    else:
                        print("No image provided or invalid image.")

                
                variants = variant_formset.save(commit=False)
                for variant in variants:
                    variant.product = product
                    variant.save()
 
                messages.success(request, "Product and variants added successfully")
                return redirect('adminProducts')
            else: 
                print("Product Form Errors:", product_form.errors)
                print("Variant Formset Errors:", variant_formset.errors)
                messages.warning(request, "Please correct the errors below.")
        else:
            product_form = ProductForm()
            variant_formset = VariantFormSet(queryset=Variant.objects.none())

        categories = Category.objects.all()
        brands = Brand.objects.all()
        return render(request, 'addproducts.html', {
            'product_form': product_form,
            'variant_formset': variant_formset,
            'categories': categories,
            'brands': brands,
        })
    return redirect(adminLogin)


@never_cache
@superuser_required
def editProducts(request, id):
    if request.user.is_superuser:
        product = get_object_or_404(Products, id=id)
        VariantFormSet = modelformset_factory(Variant, form=VariantForm, extra=0, can_delete=True)
        
        if request.method == 'POST':
            product_form = ProductForm(request.POST,instance=product)
            variant_formset = VariantFormSet(request.POST, queryset=product.variants.all())

            image_files = [
                request.POST.get('image1'),
                request.POST.get('image2'),
                request.POST.get('image3'),
                request.POST.get('image4'),
            ]

            if product_form.is_valid() and variant_formset.is_valid():
                updated_product = product_form.save()
                
                # Handle images
                existing_images = updated_product.product_images.all()
                for index, image_data in enumerate(image_files):
                    if image_data:
                        try:
                            format, imgstr = image_data.split(';base64,')
                            ext = format.split('/')[-1]
                            image_file = ContentFile(base64.b64decode(imgstr), name=f'product_image_{index+1}.{ext}')
                            if index < len(existing_images):
                                existing_image = existing_images[index]
                                existing_image.images = image_file
                                existing_image.save()
                            else:
                                Images.objects.create(images=image_file, product=updated_product)
                        except Exception as e:
                            print(f"Error processing image: {e}")
                            

                # Handle variants
                for form in variant_formset:
                    if form.cleaned_data:
                        if form.cleaned_data.get('DELETE', False):
                            if form.instance.pk:
                                form.instance.delete()
                        else:
                            form.save()

                messages.success(request, "Product and variants updated successfully.")
                return redirect('adminProducts')
            else:
                print("Product Form Errors:", product_form.errors)
                print("Variant Formset Errors:", variant_formset.errors)
                messages.warning(request, "Please correct the errors below.")
        
        else:
            product_form = ProductForm(instance=product)
            variant_formset = VariantFormSet(queryset=product.variants.all())

        return render(request, 'editproducts.html', {
            'product_form': product_form,
            'variant_formset': variant_formset,
            'product': product,
        })

    return redirect('adminLogin')


@never_cache
@superuser_required
def addVariant(request):
    if request.method == 'POST':
        form = VariantForm(request.POST)
        product_id = request.POST.get('product_id')
        if form.is_valid():
            variant = form.save(commit=False)  
            # product_id = request.POST.get('product_id') 
            try:
                product = Products.objects.get(id=product_id) 
                variant.product = product 
                variant.save()  
                messages.success(request, "Variant added successfully.")
            except Products.DoesNotExist:
                messages.error(request, "The product does not exist.")
            return redirect('adminProducts')  
        else:
            print("form is not valid")
            messages.warning(request,'The Price and Weight must be postive numbers')
            return redirect('adminProducts')
    else:
        form = VariantForm()  # Create a new form instance for a GET request

    return render(request, 'adminproducts.html', {'form': form})



@never_cache
@superuser_required
def productSearch(request):
    if request.user.is_superuser:
        if request.method=='POST':
            search = request.POST.get('search')
            if search:
                product = Products.objects.filter(name__icontains=search).order_by('name')
            else:
                product = Products.objects.all()
                
            return render(request,'adminproducts.html',{'products':product})


@never_cache
@superuser_required
def removeProducts(request,id):
    if request.user.is_superuser:
        if request.user.is_authenticated:
            if request.method=='POST':
                product = get_object_or_404(Products,id=id)
                product.is_active = not product.is_active
                product.save()
                return redirect(adminProducts)
            else:
                print("request method is not post")
        else:
            print("user is not authenticated")    
    return redirect(adminLogout)


                    #########           #########
                    #########ADMIN BRANDS########
                    #########           #########


@superuser_required
def adminBrands(request):
    if request.user.is_superuser:
        brand_list = Brand.objects.all().order_by('-updated_date')
        
        paginator = Paginator(brand_list,5)
        page_number = request.GET.get('page')
        brand = paginator.get_page(page_number)
        
        return render(request,'adminbrands.html',{'items':brand})
    return redirect(adminLogin)

@never_cache
@superuser_required
def addBrands(request):
    if request.user.is_superuser:
        if request.method=='POST':
            brand = request.POST.get('brand')
            if not brand:
                messages.warning(request,"Name cant be empty")
                return redirect(adminBrands)
            if Brand.objects.filter(name__iexact=brand).exists():
                messages.warning(request,"Brand name already exists")
                return redirect(adminBrands)
            if not re.match("^[A-Za-z ]*$", brand) or not brand.strip():
                messages.warning(request, "Name must contain only letters and cannot be empty.")
                return redirect(adminBrands)
            if len(brand) > 20:
                messages.warning(request,"Brand name must be less than 20 characters")
                return redirect(adminBrands)
            
            new_brand=Brand.objects.create(name=brand)
            new_brand.save()
            return redirect(adminBrands)
    return redirect(adminLogin)


@never_cache
@superuser_required
def removeBrands(request,id):
    if request.user.is_superuser:
        brand = Brand.objects.get(id=id)
        brand.is_active = not brand.is_active
        brand.save()
        return redirect(adminBrands)
    return redirect(adminLogin)

@never_cache
@superuser_required
def editBrands(request,id):
    if request.user.is_superuser:
        if request.method=='POST':
            new_name = request.POST.get('brand')
            if not new_name:
                messages.warning(request,"Name cant be empty")
                return redirect(adminBrands)
            if Brand.objects.filter(name__iexact=new_name).exclude(id=id).exists():
                messages.warning(request,"Brand name already exists")
                return redirect(adminBrands)
            if not re.match("^[A-Za-z ]*$", new_name) or not new_name.strip():
                messages.warning(request, "Name must contain only letters and cannot be empty.")
                return redirect(adminBrands)
            if len(new_name) > 20:
                messages.warning(request,"Brand name must be less than 20 characters")
                return redirect(adminBrands)
            
            brand = Brand.objects.get(id=id)
            brand.name=new_name
            brand.save()
            return redirect(adminBrands)
    return redirect(adminLogin)
        

@never_cache
@superuser_required
def brandSearch(request):
    if request.user.is_superuser:
        if request.method=='POST':
            search = request.POST.get('search')
            if search is not None:
                item = Brand.objects.filter(name__icontains=search).order_by('name')
            else:
                item = Brand.objects.all()
            return render(request,'adminbrands.html',{'items':item})









@superuser_required
def adminOrders(request):
    if request.user.is_superuser:     
        orders_list = OrderItem.objects.select_related('order', 'product').prefetch_related('order__user').order_by('-last_updated')
        
        paginator = Paginator(orders_list,5)
        page_number = request.GET.get('page')
        orders = paginator.get_page(page_number)
        
        return render(request, 'adminorders.html', {'orders': orders})
    else:
        return redirect(adminLogin)


@never_cache
@superuser_required
def UpdateOrderStatus(request, order_item_id):
    if request.method == 'POST':
        order_item = get_object_or_404(OrderItem, id=order_item_id)
        new_status = request.POST.get('status')
        order_item.status = new_status
        order_item.save()
        return redirect(adminOrders)


@never_cache
@superuser_required
def orderSearch(request):
    if request.user.is_superuser:
        if request.method == 'POST':
            search = request.POST.get('search')
            if search:
                result = OrderItem.objects.filter(
                    Q(product__name__icontains=search) | 
                    Q(order__user__username__icontains=search)
                ).distinct()  

                return render(request, 'adminorders.html', {'orders': result, 'search_query': search})
    return render(request, 'adminorders.html', {'orders': []})
            
                

@superuser_required
def adminCoupons(request):
    if request.user.is_superuser:      
        coupons_list  = Coupon.objects.all().order_by('-id')
        
        paginator = Paginator(coupons_list,5)
        page_number = request.GET.get('page')
        coupons = paginator.get_page(page_number) 
        
        return render(request,'admincoupons.html',{'coupons':coupons})
    return redirect(adminLogin)

@superuser_required
def addCoupon(request):
    if request.user.is_superuser:       
        if request.method == 'POST':
            form = CouponForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request,"Coupon added successfully!!!")
                return redirect('adminCoupons')  # Redirect back to the coupons list page
        else:
            form = CouponForm()
        return render(request, 'adminaddcoupons.html', {'form': form})
    return redirect(adminLogin)


@superuser_required
def editCoupon(request,coupon_id):
    if request.user.is_superuser:      
        coupon = get_object_or_404(Coupon,id=coupon_id)
        
        if request.method == "POST":
            form = CouponForm(request.POST,instance=coupon)
            if form.is_valid():
                form.save()
                messages.success(request,"Coupon edited Successfully!1")
                return redirect(adminCoupons)
        else:
            form = CouponForm(instance=coupon)
        return render(request,'admineditcoupons.html',{'form':form})
    return redirect(adminLogin)

   
@superuser_required 
def deleteCoupon(request,coupon_id):
    coupon = get_object_or_404(Coupon,id=coupon_id)
    coupon.active = not coupon.active
    coupon.save()
    messages.success(request,'Coupon Remove/Active Successfully!!')
    return redirect(adminCoupons)


@superuser_required
def searchCoupon(request):
    if request.user.is_superuser:
        if request.method == "POST":
            search = request.POST.get('search')
            if search is not None:
                coupons = Coupon.objects.filter(code__icontains=search)
            else:
                coupons = Coupon.objects.all()
            return render(request,'admincoupons.html',{'coupons':coupons})
            
        
@superuser_required
def adminOffers(request):
    if request.user.is_superuser:       
        offers_list = Offer.objects.all().order_by('-last_updated')
        
        paginator = Paginator(offers_list,5)
        page_number = request.GET.get('page')
        offers = paginator.get_page(page_number)
        
        return render(request,'adminoffers.html',{'offers':offers})
    return redirect(adminLogin)

@superuser_required
def adminAddOffers(request):
    if request.user.is_superuser:      
        if request.method == 'POST':
            form = OfferForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request,"Offers added Successfully!!!")
                return redirect(adminOffers)
            else:
                messages.warning(request,'Please correct the errors below')
                
        else:
            form = OfferForm()
        return render(request,'adminaddoffers.html',{'form':form})
    return redirect(adminLogin)


@superuser_required
def adminEditOffers(request,offer_id):
    if request.user.is_superuser:         
        try:
            offer = get_object_or_404(Offer,id=offer_id)
        except:
            messages.error(request,"Offer not found!")
            return redirect(adminOffers)
        if request.method == "POST":
            form = OfferForm(request.POST,instance=offer)
            if form.is_valid():
                offer.products.clear()
                offer.categories.clear()
                form.save()
                messages.success(request,"Offer Edited Successfully")
                return redirect(adminOffers)
            else:
                messages.warning(request,"Please correct the errors below")
        else:
            form = OfferForm(instance=offer)
        return render(request,'admineditoffers.html',{'form':form,'offer':offer})
    return redirect(adminLogin)

@superuser_required
def adminDeleteOffer(request,offer_id):
    if request.method=="POST":
        offer = get_object_or_404(Offer,id=offer_id)
        offer.is_active = not offer.is_active
        offer.save()
        return redirect(adminOffers)
    
        
        
@superuser_required
def adminSearchOffers(request):
    if request.method=="POST":
        search = request.POST.get('search')
        if search:
            offers = Offer.objects.filter(name__icontains=search)
        else:
            offers = Offer.objects.all()
        return render(request,'adminoffers.html',{'offers':offers})
   
@superuser_required   
def adminDashboard(request):
    if request.user.is_superuser:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        order = Order.objects.all()
        
        top_products = order.values('order_items__product__name').annotate(
            total_quantity=Sum('order_items__quantity'),
            total_revenue=Sum('total_amount')
        ).order_by('-total_quantity')[:10]

        # Top 10 categories
        top_categories = order.values('order_items__product__category__name').annotate(
            total_quantity=Sum('order_items__quantity'),
            total_revenue=Sum('total_amount')
        ).order_by('-total_quantity')[:10]

        # Top 10 brands
        top_brands = order.values('order_items__product__brand__name').annotate(
            total_quantity=Sum('order_items__quantity'),
            total_revenue=Sum('total_amount')
        ).order_by('-total_quantity')[:10] 
        
    
        if not start_date or not end_date:
            end_date = now()
            start_date = end_date - timedelta(days=7)
        else:
        
            start_date = make_aware(datetime.strptime(start_date, "%Y-%m-%d"))
            end_date = make_aware(datetime.strptime(end_date, "%Y-%m-%d"))
        print(start_date)
        print(end_date)
            
        
        
        orders = Order.objects.filter(created_at__range=[start_date, end_date])
        
    
        total_sales = orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        total_discounts = (
            orders.aggregate(
                total_discount=Sum(F('order_items__price') * F('order_items__quantity'), output_field=DecimalField())
            )['total_discount'] or 0
        ) - total_sales if orders.exists() else 0
        
        total_orders = orders.count()

        
        today = now().date()
        start_of_week = today - timedelta(days=today.weekday()) 
        start_of_month = today.replace(day=1) 

        
        daily_orders = Order.objects.filter(created_at__date=today)
        daily_sales = daily_orders.aggregate(total=Sum('total_amount'))['total'] or 0
        daily_discounts = (
            daily_orders.aggregate(
                discount=Sum(F('order_items__price') * F('order_items__quantity'), output_field=DecimalField())
            )['discount'] or 0
        ) - daily_sales if daily_orders.exists() else 0

        
        weekly_orders = Order.objects.filter(created_at__date__gte=start_of_week, created_at__date__lte=today)
        weekly_sales = weekly_orders.aggregate(total=Sum('total_amount'))['total'] or 0
        weekly_discounts = (
            weekly_orders.aggregate(
                discount=Sum(F('order_items__price') * F('order_items__quantity'), output_field=DecimalField())
            )['discount'] or 0
        ) - weekly_sales if weekly_orders.exists() else 0
        

    
        monthly_orders = Order.objects.filter(created_at__date__gte=start_of_month, created_at__date__lte=today)
        monthly_sales = monthly_orders.aggregate(total=Sum('total_amount'))['total'] or 0
        monthly_discounts = (
            monthly_orders.aggregate(
                discount=Sum(F('order_items__price') * F('order_items__quantity'), output_field=DecimalField())
            )['discount'] or 0
        ) - monthly_sales if monthly_orders.exists() else 0

        # chart starts here
        chart_filter = request.GET.get('filter', 'weekly')
        if chart_filter == 'weekly':
            chart_data = orders.annotate(period=TruncWeek('created_at')).values('period').annotate(
                total_sales=Sum('total_amount')
            ).order_by('period')
        elif chart_filter == 'monthly':
            chart_data = orders.annotate(period=TruncMonth('created_at')).values('period').annotate(
                total_sales=Sum('total_amount')
            ).order_by('period')
        elif chart_filter == 'yearly':
            chart_data = orders.annotate(period=TruncYear('created_at')).values('period').annotate(
                total_sales=Sum('total_amount')
            ).order_by('period')
        elif chart_filter == 'daily':
            chart_data = orders.annotate(period=TruncDay('created_at')).values('period').annotate(
                total_sales=Sum('total_amount')
            ).order_by('period')

        # Extract labels and data for the chart
        chart_labels = [data['period'].strftime('%Y-%m-%d') for data in chart_data]
        chart_values = [float(data['total_sales']) or 0 for data in chart_data]

        print("Chart Labels:", chart_labels)
        print("Chart Data:", chart_values)


        
        context = {
            'total_sales': total_sales,
            'total_discounts': total_discounts,
            'total_orders': total_orders,
            'start_date': start_date.strftime('%Y-%m-%d'), 
            'end_date': end_date.strftime('%Y-%m-%d'),  
            'daily': {'sales': daily_sales, 'discounts': daily_discounts},
            'weekly': {'sales': weekly_sales, 'discounts': weekly_discounts},
            'monthly': {'sales': monthly_sales, 'discounts': monthly_discounts},
            'top_products': top_products,
            'top_categories': top_categories,
            'top_brands': top_brands,
            'chart_labels': json.dumps(chart_labels),  # Pass as JSON
            'chart_data': json.dumps(chart_values),
        }
        
        return render(request, 'admindashboard.html', context)
    return redirect(adminLogin)


@superuser_required
def admin_dashboard_chart(request):
    filter_type = request.GET.get('filter', 'weekly')
    today = now().date()

    if filter_type == 'weekly':
        chart_data = Order.objects.annotate(period=TruncWeek('created_at')).values('period').annotate(
            total_sales=Sum('total_amount')
        ).order_by('period')
    elif filter_type == 'monthly':
        chart_data = Order.objects.annotate(period=TruncMonth('created_at')).values('period').annotate(
            total_sales=Sum('total_amount')
        ).order_by('period')
    elif filter_type == 'yearly':
        chart_data = Order.objects.annotate(period=TruncYear('created_at')).values('period').annotate(
            total_sales=Sum('total_amount')
        ).order_by('period')
    elif filter_type == 'daily' :
        chart_data = Order.objects.annotate(period=TruncDay('created_at')).values('period').annotate(
            total_sales=Sum('total_amount')
        ).order_by('period')

    labels = [data['period'].strftime('%Y-%m-%d') for data in chart_data]
    data = [float(data['total_sales']) or 0 for data in chart_data]
    
    print("Dynamic Chart Labels:", labels)
    print("Dynamic Chart Data:", data)


    return JsonResponse({'labels': labels, 'data': data})




    
@superuser_required
def export_to_excel(request):
    
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    try:
        start_date = datetime.strptime(start_date_str, "%b. %d, %Y")  
    except ValueError:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d") 

    try:
        end_date = datetime.strptime(end_date_str, "%b. %d, %Y")  
    except ValueError:
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d") 
        
    start_date = make_aware(start_date)
    end_date = make_aware(end_date)

    # Get orders within the date range
    orders = Order.objects.filter(created_at__range=[start_date, end_date])

    # Calculate total sales, total discounts, and total orders
    total_sales = sum(order.total_amount for order in orders)
    total_discounts = sum(order.total_discount for order in orders if order.total_discount)
    total_orders = len(orders)

    # Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Adding total sales, total discounts, and total orders
    ws.append([f"Total Sales: Rs.{total_sales:.2f}"])
    ws.append([f"Total Discounts: Rs.{total_discounts:.2f}"])
    ws.append([f"Total Orders: {total_orders}"])
    ws.append([])  # Empty row to separate from the headers

    # Add headers for the orders
    headers = ["Order ID", "Customer", "Total Amount", "Total Discount", "Order Date"]
    ws.append(headers)

    # Add the order details
    for order in orders:
        total_discount = sum(item.get_total_price() for item in order.order_items.all())  
        ws.append([order.id, order.user.username, order.total_amount, total_discount, order.created_at.strftime('%Y-%m-%d')])

    # Adjust column widths
    for col_num in range(1, len(headers) + 1):
        column = get_column_letter(col_num)
        max_length = 0
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Prepare the response to send the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Sales_Report_{start_date.strftime("%Y-%m-%d")}_to_{end_date.strftime("%Y-%m-%d")}.xlsx'
    
    # Save the Excel file to the response
    wb.save(response)
    return response

    



@superuser_required
def export_to_pdf(request):
   
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    try:
        start_date = datetime.strptime(start_date_str, "%b. %d, %Y")  
    except ValueError:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")  

    try:
        end_date = datetime.strptime(end_date_str, "%b. %d, %Y")  
    except ValueError:
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")  
    
    start_date = make_aware(start_date)
    end_date = make_aware(end_date)
    
    
    orders = Order.objects.filter(created_at__range=[start_date, end_date])

    
    total_sales = sum(order.total_amount for order in orders)
    total_discounts = sum(order.total_discount for order in orders if order.total_discount)
    total_orders = len(orders)

    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Sales_Report_{start_date.strftime("%Y-%m-%d")}_to_{end_date.strftime("%Y-%m-%d")}.pdf'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    
    p.setFont("Helvetica", 14)
    p.drawString(100, height - 40, f"Sales Report from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

    
    p.setFont("Helvetica", 10)
    p.drawString(100, height - 60, f"Total Sales: Rs.{total_sales:.2f}")
    p.drawString(100, height - 80, f"Total Discounts: Rs.{total_discounts:.2f}")
    p.drawString(100, height - 100, f"Total Orders: {total_orders}")

    
    p.drawString(100, height - 120, "Order ID")
    p.drawString(200, height - 120, "Customer")
    p.drawString(300, height - 120, "Total Amount")
    p.drawString(400, height - 120, "Discount Amount")
    p.drawString(500, height - 120, "Order Date")

    
    y_position = height - 140
    for order in orders:
        total_discount = sum(item.get_total_price() for item in order.order_items.all())
        p.drawString(100, y_position, str(order.id))
        p.drawString(200, y_position, order.user.username)
        p.drawString(300, y_position, f"Rs.{order.total_amount:.2f}")
        p.drawString(400, y_position, f"Rs.{total_discount:.2f}")
        p.drawString(500, y_position, order.created_at.strftime('%Y-%m-%d'))
        y_position -= 20  # Move to the next line

    
    p.showPage()
    p.save()
    
    return response



@superuser_required
def manage_requests(request):           
    if request.user.is_authenticated:
        requests = OrderItem.objects.filter(status='Processing')
        return render(request,'manage_requests.html',{'requests':requests})
     
    
    





@require_POST
@superuser_required
def approve_request(request, request_id):
    order_item = get_object_or_404(OrderItem, id=request_id, status="Processing")

    
    order = order_item.order
    
    
    total_order_price = sum(item.get_total_price() for item in order.order_items.all())
    total_discount = order.total_discount or 0

    
    if total_order_price > 0:
        discount_per_unit = total_discount / total_order_price
    else:
        discount_per_unit = 0

    
    item_total_price = order_item.get_total_price()
    discounted_refund_amount = item_total_price - (item_total_price * discount_per_unit)

    
    order_item.status = "Refunded"
    order_item.save()

    
    refund_to_wallet(
        user=order_item.order.user,
        amount=discounted_refund_amount,
        product=order_item.product,
        description=f"Refund for {order_item.product.name} (discounted)"
    )

    
    variant = Variant.objects.filter(product=order_item.product).first()
    if variant:
        variant.stock = F("stock") + order_item.quantity
        variant.save()

    messages.success(
        request, 
        f"Request for {order_item.product.name} has been approved, and a discounted amount of Rs. {discounted_refund_amount:.2f} has been refunded to the wallet."
    )
    return redirect(adminlogin)




@require_POST
@superuser_required
def reject_request(request, request_id):
    order_item = get_object_or_404(OrderItem, id=request_id, status="Processing")
    
    # Update status
    order_item.status = "Shipped"
    order_item.save()
    
    messages.success(request, f"Request for {order_item.product.name} has been rejected.")
    return redirect(manage_requests)